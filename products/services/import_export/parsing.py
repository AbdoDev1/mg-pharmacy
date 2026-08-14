"""
مرحلة القراءة والتصنيف (بدون أي حفظ) — الخطوة الأولى من استيراد إكسل.

كل الدوال هنا "نقية" قدر الإمكان: بتاخد بيانات وترجع بيانات، من غير ما
تعرف حاجة عن HTTP request أو response أو session. ده اللي بيخليها سهلة
الاختبار (products/services/tests.py) من غير ما نحتاج نمرّ بـ Django
test client أو نرفع ملف Excel فعلي لكل اختبار.

الترتيب: read_import_workbook بتقرا الملف صفًا صفًا (parse_unit_row)،
بتلمّ صفوف نفس الصنف مع بعض (group_unit_rows)، وبعدين بتحدد لكل صنف
هل ده تحديث لمنتج معروف ولا صنف جديد (classify_row). الحفظ الفعلي بعد
موافقة الموظف موجود في commit.py.
"""
from decimal import Decimal, InvalidOperation

import openpyxl

from accounts.models import AccountType
from products.matching import normalize_name, find_similar_products
from products.models import Product
from studio.models import StudioImage

from .common import FUZZY_MATCH_THRESHOLD, REQUIRED_IMPORT_HEADERS, discount_col_name

__all__ = [
    'parse_unit_row',
    'group_unit_rows',
    'classify_row',
    'read_import_workbook',
]


def parse_unit_row(row_num, row, idx, account_types_by_col):
    """
    بيقرا صف واحد من شيت الإكسل — كل صف بيمثّل وحدة واحدة بس (قطعة أو
    كرتونة) لصنف معيّن، مش الصنف كامل. الصنف اللي له وحدتين (صغرى وكبرى)
    بيتكرر في صفين بنفس code (أو نفس الاسم) — راجع group_unit_rows تحت
    اللي بتلمّهم مرة تانية في صنف واحد قبل الحفظ.
    """
    def cell(key):
        pos = idx.get(key)
        return row[pos] if pos is not None and pos < len(row) else None

    def cell_str(key):
        """مثل cell() لكن بيحوّل رقم صحيح متخزن كـ float (زي 123.0 لو الباركود
        كله أرقام وExcel اعتبره رقم) لنص صحيح "123" بدل "123.0"."""
        value = cell(key)
        if value is None:
            return ''
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    name_ar = cell_str('name_ar')
    category_slug = cell_str('category_slug')
    unit_name = cell_str('unit_name')
    # code هو الحقل الوحيد المعتمد للتفرقة بين الأصناف وقت الاستيراد —
    # الباركود مش موجود في شيت الإكسل خالص (لا قراءة ولا كتابة)، بيتسجّل
    # من صفحة المنتج نفسها فقط. راجع Product.BARCODE_FIELDS في models.py.
    code = cell_str('code')
    # عمود اختياري جديد (مرحلة 9 في STUDIO_PLAN.md) — معرّف StudioImage.pk
    # كنص خام هنا بس، بيتحقق منه فعليًا (رقم صحيح + موجود فعلًا في
    # الاستوديو) بعد التجميع في read_import_workbook تحت، مش هنا.
    studio_image_id = cell_str('studio_image_id')

    raw_qty_in_small = cell('qty_in_small')
    raw_unit_price = cell('unit_price')
    raw_quantity = cell('quantity')

    if not name_ar or not unit_name or not raw_qty_in_small or not raw_unit_price:
        return None, f'سطر {row_num}: بيانات ناقصة (الاسم/الوحدة/الكمية بالقطعة/سعر الجمهور)'

    try:
        qty_in_small = int(raw_qty_in_small)
        unit_price = round(float(raw_unit_price), 2)
        quantity = int(raw_quantity) if raw_quantity else 0
    except (TypeError, ValueError):
        return None, f'سطر {row_num}: قيم رقمية غير صالحة'

    if qty_in_small < 1:
        return None, f'سطر {row_num}: "الكمية بالوحدة الصغرى" يجب أن تكون 1 على الأقل'

    # ملحوظة: كان هنا فحص بيرفض الصف لو category_slug مكتوب في الملف
    # ومفيهوش قسم مطابق في القاعدة. اتشال عمدًا — القسم غير الموجود دلوقتي
    # بيتعمل تلقائيًا وقت الحفظ الفعلي (commit.py، عن طريق
    # get_or_create_category) بدل ما يوقّف استيراد الصف كله. هنا في مرحلة
    # القراءة إحنا بس بنتأكد إن العمود مقروء صح، مش بنتحقق من وجود القسم.

    # عمود discount:<فئة> موجود في الملف = القيمة دي هي الوصف الكامل لحالة
    # الخصم لهذا النوع، تمامًا زي شاشة "قائمة الخصومات" اليدوية: فاضي يعني
    # "امسح الخصم" مش "سيبه زي ما هو". نوع مالوش عمود أصلًا في الملف (مش
    # ضمن account_types_by_col) بيتسيب بدون أي تغيير.
    discounts = {}
    for col_name, account_type in account_types_by_col.items():
        raw = cell(col_name)
        if raw is None or str(raw).strip() == '':
            discounts[account_type.pk] = None  # إزالة صريحة لأي خصم موجود
            continue
        try:
            pct = Decimal(str(raw).strip())
        except InvalidOperation:
            return None, f'سطر {row_num}: نسبة خصم غير صالحة لـ"{account_type.name}"'
        if pct < 0 or pct > 100:
            return None, f'سطر {row_num}: نسبة الخصم لـ"{account_type.name}" يجب أن تكون بين 0 و100'
        discounts[account_type.pk] = str(pct)  # str عشان يفضل JSON-safe في الـ session

    return {
        'row_num': row_num,
        'code': code,
        'category_slug': category_slug,
        'name_ar': name_ar,
        'unit_name': unit_name,
        'qty_in_small': qty_in_small,
        'unit_price': unit_price,
        'quantity': quantity,
        'discounts': discounts,
        'studio_image_id': studio_image_id,
    }, None


def group_unit_rows(unit_rows):
    """
    بتجمع صفوف الوحدات (كل صف = وحدة واحدة) في صفوف "أصناف" — صنف بوحدة
    واحدة (كبرى بس مثلاً) أو بوحدتين (صغرى + كبرى) بيتعرف عن طريق نفس
    الـ code (لو موجود) أو نفس الاسم (بعد التطبيع) عبر كل صفوف الملف.
    نسبة الخصم بتتاخد من صف الوحدة الصغرى لو موجودة، وإلا من الوحدة
    الوحيدة (لو الصنف مالوش صغرى أصلًا) — نفس قاعدة التسعير في الموديل.
    """
    groups = {}
    order = []
    for ur in unit_rows:
        key = ('code', ur['code']) if ur['code'] else ('name', normalize_name(ur['name_ar']))
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(ur)

    products_data, errors = [], []
    for key in order:
        rows = groups[key]
        row_nums = ', '.join(str(r['row_num']) for r in rows)
        if len(rows) > 2:
            errors.append(f'صنف "{rows[0]["name_ar"]}": أكتر من وحدتين للصنف ده في الملف (سطور {row_nums})')
            continue
        small_rows = [r for r in rows if r['qty_in_small'] == 1]
        large_rows = [r for r in rows if r['qty_in_small'] > 1]
        if len(small_rows) > 1 or len(large_rows) > 1:
            errors.append(f'صنف "{rows[0]["name_ar"]}": فيه أكتر من صف بنفس حجم الوحدة (سطور {row_nums})')
            continue
        small = small_rows[0] if small_rows else None
        large = large_rows[0] if large_rows else None
        discount_source = small or large
        category_slug = next((r['category_slug'] for r in rows if r['category_slug']), '')
        code = next((r['code'] for r in rows if r['code']), '')
        # .get() مش ['...'] عمدًا هنا (بخلاف category_slug/code فوق) —
        # اختبارات موجودة قبل مرحلة 9 بتبني صفوف الوحدة يدويًا كـ dict
        # بلا المفتاح ده خالص، فالوصول المباشر كان بيطلع KeyError عليها.
        studio_image_id = next((r.get('studio_image_id') for r in rows if r.get('studio_image_id')), '')
        products_data.append({
            'row_num': rows[0]['row_num'],
            'row_nums': [r['row_num'] for r in rows],
            'code': code,
            'category_slug': category_slug,
            'name_ar': rows[0]['name_ar'],
            'small': small,
            'large': large,
            'discounts': discount_source['discounts'] if discount_source else {},
            'studio_image_id': studio_image_id,
        })
    return products_data, errors


def classify_row(row_data, existing_by_code, existing_by_name_key, all_products):
    """
    بيحدد إيه اللي المفروض يحصل للصف ده اعتمادًا على مطابقة الكود (لو
    موجود) ثم الاسم المُطبَّع (بعد إزالة فروق المسافات/الأرقام/الحروف
    الشكلية) ثم أقرب الأسماء تشابهًا لو مفيش تطابق تام. النتيجة action
    واحدة من: update (تحديث صنف معروف بثقة) أو review (يحتاج قرار بشري)
    أو create (صنف جديد فعلًا، مفيش أي شبه بحاجة موجودة).
    """
    name_key = normalize_name(row_data['name_ar'])
    row_data['name_key'] = name_key

    if row_data['code'] and row_data['code'] in existing_by_code:
        product = existing_by_code[row_data['code']]
        return {**row_data, 'action': 'update', 'match_pk': product.pk,
                'match_name': product.name_ar, 'match_reason': 'code'}

    if name_key in existing_by_name_key:
        product = existing_by_name_key[name_key]
        return {**row_data, 'action': 'update', 'match_pk': product.pk,
                'match_name': product.name_ar, 'match_reason': 'name'}

    candidates = find_similar_products(name_key, all_products, threshold=FUZZY_MATCH_THRESHOLD)
    if candidates:
        return {**row_data, 'action': 'review', 'match_pk': None,
                'candidates': [{'pk': p.pk, 'name': p.name_ar, 'code': p.code, 'score': s}
                                for p, s in candidates]}

    return {**row_data, 'action': 'create', 'match_pk': None}


def read_import_workbook(excel_file, max_rows):
    """
    بتقرا ملف Excel كامل وترجّع (rows, errors, error_message).
    error_message لو موجودة معناها فشل عام (ملف غير صالح، أعمدة ناقصة،
    عدد صفوف أكبر من الحد المسموح) والعملية لازم توقف فورًا. غير كده،
    rows هي قائمة الأصناف المصنّفة (update/create/review) وerrors تحذيرات
    على مستوى صف واحد بس (باقي الصفوف الصحيحة اتعالجت عادي).
    """
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        missing = [h for h in REQUIRED_IMPORT_HEADERS if h not in headers]
        if missing:
            return [], [], f'الأعمدة التالية ناقصة في الملف: {", ".join(missing)}'
        idx = {h: headers.index(h) for h in headers if h}

        account_types = list(AccountType.objects.all().order_by('name'))
        account_types_by_col = {
            discount_col_name(at): at for at in account_types if discount_col_name(at) in idx
        }

        all_products = list(Product.objects.only('id', 'name_ar', 'code', 'name_key'))
        existing_by_code = {p.code: p for p in all_products if p.code}
        existing_by_name_key = {p.name_key: p for p in all_products if p.name_key}

        unit_rows, errors = [], []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            if row_num - 1 > max_rows:
                wb.close()
                return [], [], f'الملف فيه أكتر من {max_rows} صف. يرجى تقسيمه على أكتر من دفعة استيراد.'
            try:
                row_data, error = parse_unit_row(row_num, row, idx, account_types_by_col)
            except Exception as e:
                error = f'سطر {row_num}: خطأ — {str(e)}'
                row_data = None
            if error:
                errors.append(error)
                continue
            unit_rows.append(row_data)

        wb.close()
    except Exception as e:
        return [], [], f'خطأ في قراءة الملف: {str(e)}'

    products_data, group_errors = group_unit_rows(unit_rows)
    errors.extend(group_errors)

    # تحقق معرّف صورة الاستوديو الاختياري (مرحلة 9 في STUDIO_PLAN.md) —
    # بعد التجميع (مش وقت parse_unit_row) عشان نتحقق مرة واحدة بس لكل
    # صنف (مش لكل صف وحدة)، وبنستعلم StudioImage مرة واحدة بكل الـ IDs
    # المطلوبة دفعة واحدة بدل استعلام منفصل لكل صنف. عمود فاضي = مفيش
    # تغيير مطلوب على صورة الصنف (مش خطأ، ومفيش تحذير) — التحذير بس لو
    # اتكتب معرّف فعليًا ومش رقم صحيح أو مش موجود في الاستوديو خالص.
    requested_ids = set()
    for p in products_data:
        raw = p.get('studio_image_id', '')
        if raw:
            try:
                requested_ids.add(int(raw))
            except (TypeError, ValueError):
                pass
    existing_image_ids = set(
        StudioImage.objects.filter(pk__in=requested_ids).values_list('pk', flat=True)
    ) if requested_ids else set()

    for p in products_data:
        raw = p.get('studio_image_id', '')
        if not raw:
            p['studio_image_id'] = None
            continue
        try:
            image_id = int(raw)
        except (TypeError, ValueError):
            errors.append(
                f'سطر {p["row_num"]}: معرّف صورة الاستوديو "{raw}" غير صالح — '
                f'تم تجاهل الصورة لصنف "{p["name_ar"]}"'
            )
            p['studio_image_id'] = None
            continue
        if image_id not in existing_image_ids:
            errors.append(
                f'سطر {p["row_num"]}: لا توجد صورة استوديو بمعرّف {image_id} — '
                f'تم تجاهل الصورة لصنف "{p["name_ar"]}"'
            )
            p['studio_image_id'] = None
        else:
            p['studio_image_id'] = image_id

    rows = [
        classify_row(p, existing_by_code, existing_by_name_key, all_products)
        for p in products_data
    ]

    # صنف جديد (مش تحديث لصنف معروف) لازم يكون له قسم محدد في الملف —
    # صنف بيتحدّث (update) ممكن يسيب category_slug فاضي ويفضل على قسمه الحالي.
    valid_rows = []
    for r in rows:
        if r['action'] == 'create' and not r['category_slug']:
            errors.append(f'سطر {r["row_num"]}: صنف جديد "{r["name_ar"]}" لازم يكون له قسم (category_slug)')
            continue
        valid_rows.append(r)

    return valid_rows, errors, None
